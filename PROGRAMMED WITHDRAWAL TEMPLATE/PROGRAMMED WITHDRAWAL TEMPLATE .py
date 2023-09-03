'''Importing all modules for use'''
from openpyxl import Workbook, load_workbook
import numpy_financial as np
from openpyxl.styles import numbers
import math
import fontstyle

option = input("Enter A for Annuity or P for Programmed Withdrawal: ").lower()
while True:

    if option.lower() == "a":
        print()
        print("ANNUITY QUOTE COMPUTATION")
        print()
        rsa = float(input("Enter current RSA Balance: "))
        buffer = float(input("Enter Retiree monthly PW: "))
        quote = rsa - buffer

        print(f"Retiree Annuity Quote is: {quote:,.2f}")
        quit()
    elif option.lower() == "p":
        break
    else:
        print("Invalid entry. Please Enter A for Annuity or P for Programmed Withdrawal")
        option = input("Enter A for Annuity or P for Programmed Withdrawal: ")
        continue

print("PROGRAMMED WITHDRAWAL COMPUTATION")
print()

retiree_name = input("Enter Retiree's name in capital letters as it appears on Cpas: ").upper()
retiree_pin = input("Enter Retiree PIN number: ")
compute = input(f"Do you want to compute ATE for {retiree_name} (YES/NO)? ").lower()
while compute.lower() != "no":
    if compute.lower() == "yes":
        pass
    else:
        print("Invalid entry. Please enter (YES/N0).")
        compute = input(f"Do you want to compute ATE for {retiree_name} (YES/NO)? ").lower()
        continue

    with open("ATE OPTION.txt","r") as ate:
        x = ate.read()
        print()
        print(x)

    '''Accessing the PenCom Excel Template'''
    wb = load_workbook("TEMPLATE.xlsx",data_only=True)
    ws_variables = wb["Variables Template"]
    ws_male = wb["Male"]
    ws_female = wb["Female"]

    ''''Printing Empty space'''
    print()

    ''' Taking retiree information and calculating cell values'''
    user = int(input("Enter 1,2,3,4,5,6,7 or 8 for the option above: "))
    c11 = int(input("Enter 1 for Male or 0 for Female: "))
    c12 = float(input("Enter RSA balance as at consolidated date: "))
    dob = int(input("Enter year of Birth: "))
    dor = int(input("Enter year of Retirement: "))
    c14 = dor - dob
    if c14 < 50:
        print("Retiree age is less than 50. Retiree not eligible.")
        break
    else:
        pass
    c15 = 0
    c16 = 0.25 * c12
    c19 = 0.05
    c20 = 0.003
    c21 = 0.08
    c22 = c21 * (1 - (c19 + c20))
    c25 = 12


    ''''Writing formulae for cell 23'''
    def computation():
        if (c11 == 1 and c14 == (ws_male["B37"].value)):
            return (ws_male["L37"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B38"].value)):
            return (ws_male["L38"].value)
        elif (c11 == 1 and c14 == (ws_male["B39"].value)):
            return (ws_male["L39"].value)
        elif (c11 ==1 and c14) == (ws_male["B40"].value):
            return (ws_male["40"].value)
        elif (c11 ==1 and c14 == (ws_male["B41"].value)):
            return (ws_male["B41"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B42"].value)):
            return (ws_male["L42"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B43"].value)):
            return (ws_male["L43"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B44"].value)):
            return (ws_male["L44"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B45"].value)):
            return (ws_male["L45"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B46"].value)):
            return (ws_male["L46"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B47"].value)):
            return (ws_male["L47"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B48"].value)):
            return (ws_male["L48"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B49"].value)):
            return (ws_male["L49"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B50"].value)):
            return (ws_male["L50"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B51"].value)):
            return (ws_male["L51"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B52"].value)):
            return (ws_male["L52"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B53"].value)):
            return (ws_male["L53"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B54"].value)):
            return (ws_male["L54"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B55"].value)):
            return (ws_male["L55"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B56"].value)):
            return (ws_male["L56"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B57"].value)):
            return (ws_male["L57"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B58"].value)):
            return (ws_male["L58"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B59"].value)):
            return (ws_male["L59"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B60"].value)):
            return (ws_male["L60"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B61"].value)):
            return (ws_male["L61"].value)
        elif (c11 == 1 and c14 ==  (ws_male["B62"].value)):
            return (ws_male["L62"].value)

        elif (c11 == 0 and c14 == (ws_female["B37"].value)):
            return (ws_female["B37"].value)
        elif (c11 == 0 and c14 == (ws_female["B38"].value)):
            return (ws_female["B38"].value)
        elif (c11 == 0 and c14 == (ws_female["B39"].value)):
            return (ws_female["B39"].value)
        elif (c11 == 0 and c14 == (ws_female["B40"].value)):
            return (ws_female["B40"].value)
        elif (c11 == 0 and c14 == (ws_female["B41"].value)):
            return (ws_female["B41"].value)
        elif (c11 == 0 and c14 == (ws_female["B42"].value)):
            return (ws_female["B42"].value)
        elif (c11 == 0 and c14 == (ws_female["B43"].value)):
            return (ws_female["B43"].value)
        elif (c11 == 0 and c14 == (ws_female["B44"].value)):
            return (ws_female["B44"].value)
        elif (c11 == 0 and c14 == (ws_female["B45"].value)):
            return (ws_female["B45"].value)
        elif (c11 == 0 and c14 == (ws_female["B46"].value)):
            return (ws_female["B46"].value)
        elif (c11 == 0 and c14 == (ws_female["B47"].value)):
            return (ws_female["B47"].value)
        elif (c11 == 0 and c14 == (ws_female["B48"].value)):
            return (ws_female["B48"].value)
        elif (c11 == 0 and c14 == (ws_female["B49"].value)):
            return (ws_female["B49"].value)
        elif (c11 == 0 and c14 == (ws_female["B50"].value)):
            return (ws_female["B50"].value)
        elif (c11 == 0 and c14 == (ws_female["B51"].value)):
            return (ws_female["B51"].value)
        elif (c11 == 0 and c14 == (ws_female["B52"].value)):
            return (ws_female["B52"].value)
        elif (c11 == 0 and c14 == (ws_female["B53"].value)):
            return (ws_female["B53"].value)
        elif (c11 == 0 and c14 == (ws_female["B54"].value)):
            return (ws_female["B54"].value)
        elif (c11 == 0 and c14 == (ws_female["B55"].value)):
            return (ws_female["B55"].value)
        elif (c11 == 0 and c14 == (ws_female["B55"].value)):
            return (ws_female["B55"].value)
        elif (c11 == 0 and c14 == (ws_female["B56"].value)):
            return (ws_female["B56"].value)
        elif (c11 == 0 and c14 == (ws_female["B57"].value)):
            return (ws_female["B57"].value)
        elif (c11 == 0 and c14 == (ws_female["B58"].value)):
            return (ws_female["B58"].value)
        elif (c11 == 0 and c14 == (ws_female["B59"].value)):
            return (ws_female["B59"].value)
        elif (c11 == 0 and c14 == (ws_female["B60"].value)):
            return (ws_female["B60"].value)
        elif (c11 == 0 and c14 == (ws_female["B61"].value)):
            return (ws_female["B61"].value)
        elif (c11 == 0 and c14 == (ws_female["B62"].value)):
            return (ws_female["B62"].value)
        else:
            return ("Invalid entry, try Agan!")



    '''Final salary computation'''
    def option_1():
        basic_salary = float(input("Enter BASIC SALARY: "))
        housing_alowance = float(input("Enter HOUSING ALLOWANCE: "))
        transport_allowance = float(input("Enter TRANSPORT ALLOWANCE: "))
        final_salary = (basic_salary + housing_alowance + transport_allowance) * 12
        return final_salary

    def option_2():
        basic_salary = float(input("Enter BASIC SALARY: "))
        final_salary = ((basic_salary) * 12) / 2
        return final_salary

    def option_3():
        basic_salary = float(input("Enter BASIC SALARY: "))
        housing_allowance = float(input("Enter HOUSING ALLOWANCE: "))
        final_salary = ((basic_salary + housing_allowance) * 12) / 2
        return final_salary

    def option_4():
        basic_salary = float(input("Enter BASIC SALARY: "))
        transport_allowance = float(input("Enter TRANSPORT ALLOWANCE: "))
        final_salary = ((basic_salary + transport_allowance) * 12) / 2
        return final_salary

    def option_5():
        consolidated_salary = float(input("Enter CONSOLIDATED SALARY: "))
        housing_allowance = float(input("Enter HOUSING ALLOWANCE: "))
        transport_allowance = float(input("Enter TRANSPORT ALLOWANCE: "))
        final_salary = (consolidated_salary + housing_allowance + transport_allowance) * 12
        return final_salary

    def option_6():
        consolidated_salary = float(input("Enter BASIC SALARY: "))
        housing_allowance = float(input("Enter HOUSING ALLOWANCE: "))
        final_salary = ((consolidated_salary + housing_allowance) * 12) / 2
        return final_salary

    def option_7():
        user2 = float(input("Enter CONSOLIDATED SALARY: "))
        final_salary = ((user2) * 12) / 2
        return final_salary

    def option_8():
        consolidated_salary = float(input("Enter BASIC SALARY: "))
        transport_allowance = float(input("Enter HOUSING ALLOWANCE: "))
        final_salary = ((consolidated_salary + transport_allowance) * 12) / 2
        return final_salary


    if user == 1:
        c13 = option_1()

    elif user == 2:
        c13 = option_2()

    if user == 3:
        c13 = option_3()

    elif user == 4:
        c13 = option_4()

    elif user == 5:
        c13 = option_5()

    elif user == 6:
        c13 = option_6()

    elif user == 7:
        c13 = option_7()

    elif user == 8:
        c13 = option_8()

    c23 = computation()
    c24 = c23 - (11/24)
    pmt = np.pmt(c22 / 12, 2 * c24 * c25, c12 - c15, 0, 1)
    c27 = -1 * pmt
    c26 = (c13 / 12) * 0.5
    pv = np.pv(c22 / 12, 2 * c24 * c25, c26, 0, 1)
    c17 = max(0,(c12 + pv))

    def rlswa():
        if (c17 < c16):
            return c16
        elif (c17 > (c12 / 2)):
            return (c12 / 2)
        else:
            return c17

    c18 = rlswa()

    def rmdda():
        if (c18 < c15):
            return "Error"
        elif ((c17 > c16) and (c18 > c17)):
            return "Error"
        elif ((c17 < c16) and (c18 > c16)):
            return "Error"
        else:
            var = np.pmt((c22/12),(2 * c24 * c25), (c12 - c18), 0, 1)
            var2 = -1 * var
            return var2
    c28 = rmdda()

    mc = 100 * c19
    rc = 100 * c20
    ir = 100 * c21
    irn = 100 * c22

    print()
    print(50*" ",fontstyle.apply(retiree_name,"bold/BLUE_BG"))
    print(50*" ",fontstyle.apply(retiree_pin,"bold/BLUE_BG"))
    print("Male (1)/Female(0)" + 50*"_" + "{0:,}".format(c11))
    print("RSA Balance" + 48*"_" + "{0:,.0f}".format(c12))
    print("Final Salary" + 48*"_" + "{0:,.0f}".format(c13))
    print("Age at Retirement" + 50*"_" + "{0:,}".format(c14))
    print("Min. Lump Sum Withdrawal" + 44*"_" + "{0:,}".format(c15))
    print("25% Lump Sum" + 44*"_" + "{0:,.2f}".format(c16))
    print("Max. Statutory Lump Sum Withdrawal" + 23*"_" + "{0:,.2f}".format(c17))
    print(fontstyle.apply("Recommended Lump Sum Withdrawal Amount" + 19*"_" + "{0:,.2f}".format(c18),"bold/BLUE_BG"))
    print("Management Charges" + 46*"_" + "{0:.2f}%".format(mc))
    print("Regulatory Charges" + 46*"_" + "{0:.2f}%".format(rc))
    print("Interest Rate" + 51*"_" + "{0:.2f}%".format(ir))
    print("Interest Rate Net of Charges" + 36*"_" + "{0:.2f}%".format(irn))
    print("Nx/Dx" + 53*"_" + "{0:.8f}".format(c23))
    print("nc" + 58*"_" + "{0:.6f}".format(c24))
    print("Frequency of Withdrawal/Annum" + 38*"_" + "{}".format(c25))
    print("Min. Statutory Monthly Draw down" + 28*"_" + "{0:,.2f}".format(c26))
    print("Max. Monthly Draw Down" + 37*"_" + "{0:,.2f}".format(c27))
    print(fontstyle.apply("Recommended Monthly Draw Down Amount" + 24*"_" + "{0:,.2f}".format(c28),"bold/BLUE_BG"))

    print()

    fresh_annuity = input(f"Do you want to compute annuity plan for {retiree_name} (YES/NO)? ").lower()
    if fresh_annuity.lower() == "yes":
        pw = int(input("Enter month of arears: "))
        arrears = pw * c28
        lump_sum = c18
        balance = c12
        annuity_quote = balance - (lump_sum + arrears)

        print(f"{retiree_name} Annuity Quote is {annuity_quote:,.2f}")


    else:
        pass


    compute = input(f"Do you want to recompute ATE for {retiree_name} (YES/NO)? ").lower()
    while compute.lower() != "no":
        if compute.lower() == "yes":
            choice = input(f"What is {retiree_name} Recommended Lump Sum choice?. Enter A for Min. Lump Sum,\n B for 25% Lump Sum, C for Max. Statutory Lump Sum or D to Quit? ").lower()

            def retiree_choice():
                if choice == "a":
                    return c15
                elif choice == "b":
                    return c16
                elif choice == "c":
                    return c17
                else:
                    quit()


            with open("ATE OPTION.txt", "r") as ate:
                x = ate.read()
                print()
                print(x)

            wb = load_workbook("TEMPLATE.xlsx", data_only=True)
            ws_variables = wb["Variables Template"]
            ws_male = wb["Male"]
            ws_female = wb["Female"]

            print()
            user = int(input("Enter 1,2,3,4,5,6,7 or 8 for the option above: "))
            c11 = int(input("Enter 1 for Male or 0 for Female: "))
            c12 = float(input("Enter RSA balance as at consolidated date: "))
            dob = int(input("Enter year of Birth: "))
            dor = int(input("Enter year of Retirement: "))
            c14 = dor - dob
            if c14 < 50:
                print("Retiree age is less than 50. Retiree not eligible.")
                break
            else:
                pass
            c15 = 0
            c16 = 0.25 * c12
            c19 = 0.05
            c20 = 0.003
            c21 = 0.08
            c22 = c21 * (1 - (c19 + c20))
            c25 = 12


            def computation():
                if (c11 == 1 and c14 == (ws_male["B37"].value)):
                    return (ws_male["L37"].value)
                elif (c11 == 1 and c14 == (ws_male["B38"].value)):
                    return (ws_male["L38"].value)
                elif (c11 == 1 and c14 == (ws_male["B39"].value)):
                    return (ws_male["L39"].value)
                elif (c11 == 1 and c14) == (ws_male["B40"].value):
                    return (ws_male["40"].value)
                elif (c11 == 1 and c14 == (ws_male["B41"].value)):
                    return (ws_male["B41"].value)
                elif (c11 == 1 and c14 == (ws_male["B42"].value)):
                    return (ws_male["L42"].value)
                elif (c11 == 1 and c14 == (ws_male["B43"].value)):
                    return (ws_male["L43"].value)
                elif (c11 == 1 and c14 == (ws_male["B44"].value)):
                    return (ws_male["L44"].value)
                elif (c11 == 1 and c14 == (ws_male["B45"].value)):
                    return (ws_male["L45"].value)
                elif (c11 == 1 and c14 == (ws_male["B46"].value)):
                    return (ws_male["L46"].value)
                elif (c11 == 1 and c14 == (ws_male["B47"].value)):
                    return (ws_male["L47"].value)
                elif (c11 == 1 and c14 == (ws_male["B48"].value)):
                    return (ws_male["L48"].value)
                elif (c11 == 1 and c14 == (ws_male["B49"].value)):
                    return (ws_male["L49"].value)
                elif (c11 == 1 and c14 == (ws_male["B50"].value)):
                    return (ws_male["L50"].value)
                elif (c11 == 1 and c14 == (ws_male["B51"].value)):
                    return (ws_male["L51"].value)
                elif (c11 == 1 and c14 == (ws_male["B52"].value)):
                    return (ws_male["L52"].value)
                elif (c11 == 1 and c14 == (ws_male["B53"].value)):
                    return (ws_male["L53"].value)
                elif (c11 == 1 and c14 == (ws_male["B54"].value)):
                    return (ws_male["L54"].value)
                elif (c11 == 1 and c14 == (ws_male["B55"].value)):
                    return (ws_male["L55"].value)
                elif (c11 == 1 and c14 == (ws_male["B56"].value)):
                    return (ws_male["L56"].value)
                elif (c11 == 1 and c14 == (ws_male["B57"].value)):
                    return (ws_male["L57"].value)
                elif (c11 == 1 and c14 == (ws_male["B58"].value)):
                    return (ws_male["L58"].value)
                elif (c11 == 1 and c14 == (ws_male["B59"].value)):
                    return (ws_male["L59"].value)
                elif (c11 == 1 and c14 == (ws_male["B60"].value)):
                    return (ws_male["L60"].value)
                elif (c11 == 1 and c14 == (ws_male["B61"].value)):
                    return (ws_male["L61"].value)
                elif (c11 == 1 and c14 == (ws_male["B62"].value)):
                    return (ws_male["L62"].value)

                elif (c11 == 0 and c14 == (ws_female["B37"].value)):
                    return (ws_female["B37"].value)
                elif (c11 == 0 and c14 == (ws_female["B38"].value)):
                    return (ws_female["B38"].value)
                elif (c11 == 0 and c14 == (ws_female["B39"].value)):
                    return (ws_female["B39"].value)
                elif (c11 == 0 and c14 == (ws_female["B40"].value)):
                    return (ws_female["B40"].value)
                elif (c11 == 0 and c14 == (ws_female["B41"].value)):
                    return (ws_female["B41"].value)
                elif (c11 == 0 and c14 == (ws_female["B42"].value)):
                    return (ws_female["B42"].value)
                elif (c11 == 0 and c14 == (ws_female["B43"].value)):
                    return (ws_female["B43"].value)
                elif (c11 == 0 and c14 == (ws_female["B44"].value)):
                    return (ws_female["B44"].value)
                elif (c11 == 0 and c14 == (ws_female["B45"].value)):
                    return (ws_female["B45"].value)
                elif (c11 == 0 and c14 == (ws_female["B46"].value)):
                    return (ws_female["B46"].value)
                elif (c11 == 0 and c14 == (ws_female["B47"].value)):
                    return (ws_female["B47"].value)
                elif (c11 == 0 and c14 == (ws_female["B48"].value)):
                    return (ws_female["B48"].value)
                elif (c11 == 0 and c14 == (ws_female["B49"].value)):
                    return (ws_female["B49"].value)
                elif (c11 == 0 and c14 == (ws_female["B50"].value)):
                    return (ws_female["B50"].value)
                elif (c11 == 0 and c14 == (ws_female["B51"].value)):
                    return (ws_female["B51"].value)
                elif (c11 == 0 and c14 == (ws_female["B52"].value)):
                    return (ws_female["B52"].value)
                elif (c11 == 0 and c14 == (ws_female["B53"].value)):
                    return (ws_female["B53"].value)
                elif (c11 == 0 and c14 == (ws_female["B54"].value)):
                    return (ws_female["B54"].value)
                elif (c11 == 0 and c14 == (ws_female["B55"].value)):
                    return (ws_female["B55"].value)
                elif (c11 == 0 and c14 == (ws_female["B55"].value)):
                    return (ws_female["B55"].value)
                elif (c11 == 0 and c14 == (ws_female["B56"].value)):
                    return (ws_female["B56"].value)
                elif (c11 == 0 and c14 == (ws_female["B57"].value)):
                    return (ws_female["B57"].value)
                elif (c11 == 0 and c14 == (ws_female["B58"].value)):
                    return (ws_female["B58"].value)
                elif (c11 == 0 and c14 == (ws_female["B59"].value)):
                    return (ws_female["B59"].value)
                elif (c11 == 0 and c14 == (ws_female["B60"].value)):
                    return (ws_female["B60"].value)
                elif (c11 == 0 and c14 == (ws_female["B61"].value)):
                    return (ws_female["B61"].value)
                elif (c11 == 0 and c14 == (ws_female["B62"].value)):
                    return (ws_female["B62"].value)
                else:
                    return ("Invalid entry, try Agan!")



            def option_1():
                basic_salary = float(input("Enter BASIC SALARY: "))
                housing_alowance = float(input("Enter HOUSING ALLOWANCE: "))
                transport_allowance = float(input("Enter TRANSPORT ALLOWANCE: "))
                final_salary = (basic_salary + housing_alowance + transport_allowance) * 12
                return final_salary


            def option_2():
                basic_salary = float(input("Enter BASIC SALARY: "))
                final_salary = ((basic_salary) * 12) / 2
                return final_salary


            def option_3():
                basic_salary = float(input("Enter BASIC SALARY: "))
                housing_allowance = float(input("Enter HOUSING ALLOWANCE: "))
                final_salary = ((basic_salary + housing_allowance) * 12) / 2
                return final_salary


            def option_4():
                basic_salary = float(input("Enter BASIC SALARY: "))
                transport_allowance = float(input("Enter TRANSPORT ALLOWANCE: "))
                final_salary = ((basic_salary + transport_allowance) * 12) / 2
                return final_salary


            def option_5():
                consolidated_salary = float(input("Enter CONSOLIDATED SALARY: "))
                housing_allowance = float(input("Enter HOUSING ALLOWANCE: "))
                transport_allowance = float(input("Enter TRANSPORT ALLOWANCE: "))
                final_salary = (consolidated_salary + housing_allowance + transport_allowance) * 12
                return final_salary


            def option_6():
                consolidated_salary = float(input("Enter BASIC SALARY: "))
                housing_allowance = float(input("Enter HOUSING ALLOWANCE: "))
                final_salary = ((consolidated_salary + housing_allowance) * 12) / 2
                return final_salary


            def option_7():
                user2 = float(input("Enter CONSOLIDATED SALARY: "))
                final_salary = ((user2) * 12) / 2
                return final_salary


            def option_8():
                consolidated_salary = float(input("Enter BASIC SALARY: "))
                transport_allowance = float(input("Enter HOUSING ALLOWANCE: "))
                final_salary = ((consolidated_salary + transport_allowance) * 12) / 2
                return final_salary


            if user == 1:
                c13 = option_1()

            elif user == 2:
                c13 = option_2()

            if user == 3:
                c13 = option_3()

            elif user == 4:
                c13 = option_4()

            elif user == 5:
                c13 = option_5()

            elif user == 6:
                c13 = option_6()

            elif user == 7:
                c13 = option_7()

            elif user == 8:
                c13 = option_8()

            c23 = computation()
            c24 = c23 - (11 / 24)
            pmt = np.pmt(c22 / 12, 2 * c24 * c25, c12 - c15, 0, 1)
            c27 = -1 * pmt
            c26 = (c13 / 12) * 0.5
            pv = np.pv(c22 / 12, 2 * c24 * c25, c26, 0, 1)
            c17 = max(0, (c12 + pv))

            def rlswa():
                if (c17 < c16):
                    return c16
                else:
                    return c17


            c18 = retiree_choice()

            def rmdda():
                if (c18 < c15):
                    return "Error"
                elif ((c17 > c16) and (c18 > c17)):
                    return "Error"
                elif ((c17 < c16) and (c18 > c16)):
                    return "Error"
                else:
                    var = np.pmt((c22 / 12), (2 * c24 * c25), (c12 - c18), 0, 1)
                    var2 = -1 * var
                    return var2


            c28 = rmdda()

            mc = 100 * c19
            rc = 100 * c20
            ir = 100 * c21
            irn = 100 * c22

            print()
            print(50 * " ", fontstyle.apply(retiree_name, "bold/BLUE_BG"))
            print(50 * " ", fontstyle.apply(retiree_pin, "bold/BLUE_BG"))
            print("Male (1)/Female(0)" + 50 * "_" + "{0:,}".format(c11))
            print("RSA Balance" + 48 * "_" + "{0:,.0f}".format(c12))
            print("Final Salary" + 48 * "_" + "{0:,.0f}".format(c13))
            print("Age at Retirement" + 50 * "_" + "{0:,}".format(c14))
            print("Min. Lump Sum Withdrawal" + 44 * "_" + "{0:,}".format(c15))
            print("25% Lump Sum" + 44 * "_" + "{0:,.2f}".format(c16))
            print("Max. Statutory Lump Sum Withdrawal" + 23 * "_" + "{0:,.2f}".format(c17))
            print(fontstyle.apply("Recommended Lump Sum Withdrawal Amount" + 19*"_" + "{0:,.2f}".format(c18),"bold/BLUE_BG"))
            print("Management Charges" + 46 * "_" + "{0:.2f}%".format(mc))
            print("Regulatory Charges" + 46 * "_" + "{0:.2f}%".format(rc))
            print("Interest Rate" + 51 * "_" + "{0:.2f}%".format(ir))
            print("Interest Rate Net of Charges" + 36 * "_" + "{0:.2f}%".format(irn))
            print("Nx/Dx" + 53 * "_" + "{0:.8f}".format(c23))
            print("nc" + 58 * "_" + "{0:.6f}".format(c24))
            print("Frequency of Withdrawal/Annum" + 38 * "_" + "{}".format(c25))
            print("Min. Statutory Monthly Draw down" + 28 * "_" + "{0:,.2f}".format(c26))
            print("Max. Monthly Draw Down" + 37 * "_" + "{0:,.2f}".format(c27))
            print(fontstyle.apply("Recommended Monthly Draw Down Amount" + 24*"_" + "{0:,.2f}".format(c28),"bold/BLUE_BG"))

            print()

            fresh_annuity = input(f"Do you want to compute annuity plan for {retiree_name} (YES/NO)? ").lower()
            print()
            if fresh_annuity.lower() == "yes":
                arrears = int(input("Enter month of arears: "))
                lump_sum = c18
                balance = c12
                annuity_quote = balance - (lump_sum - arrears)

                print(f"{retiree_name} Annuity Quote is {annuity_quote:,.2f}")
                print()

            else:
                pass

        compute = input(f"Do you want to recompute ATE for {retiree_name} (YES/NO)? ").lower()


    print("Thank you for using DAVE.")


print("Bye!")
